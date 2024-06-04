# Moving our data from our CRM webscraper into our existing data storage files

import shutil
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta


# getting the dates for the source file name
DateAfterMonth = datetime.today()+ relativedelta(months=1)
dateformat = DateAfterMonth.strftime("%-m-%-d-%Y")

BackTrackDate = datetime.today()- timedelta(89)
dateformat2 = BackTrackDate.strftime("%-m-%-d-%Y")

# convert to csv to remove error
shutil.move("/Users/oscarbown/Downloads/Series Expirations " +dateformat2+ " - " +dateformat+ ".xls", '/Users/oscarbown/Desktop/The Font/Python automation/Monthly membership Converted.csv')

import csv
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

with open('/Users/oscarbown/Desktop/The Font/Python automation/Monthly membership Converted.csv') as f:
    reader = csv.reader(f, delimiter=':')
    for row in reader:
        ws.append(row)

wb.save('/Users/oscarbown/Desktop/The Font/Python automation/Monthly membership Converted2.xlsx')


# moving data to Tableau file
import openpyxl as xl;
  
# opening the source excel file
filename = "/Users/oscarbown/Desktop/The Font/Python automation/Monthly membership Converted2.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]
  
# opening the destination excel file 
filename1 ="/Users/oscarbown/Desktop/The Font/RAW DATA for Tableau/Monthly membership- date & No. of visits.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active
  
# calculate total number of rows and 
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column
  
# copying the cell values from source 
# excel file to destination excel file
for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws2.cell(row = i, column = j).value = c.value
  
# saving the destination excel file
wb2.save(str(filename1))
